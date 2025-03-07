"""
jira_issues_with_sprints.py
A script to handle various transitions and operations related to data processing.
This script utilizes several libraries to fetch, manipulate, and save data as needed.
Usage:
    python jira_issues_with_sprints.py
"""
import os
import sys
import argparse
from datetime import datetime
import subprocess
import math
import re
import requests
from requests.auth import HTTPBasicAuth
from requests.exceptions import RequestException
from dotenv import load_dotenv
import urllib3
from openpyxl import Workbook
import pytz  # To handle timezone conversion
from dateutil import parser


def validate_environment_variables():
    """Validate that the necessary environment variables are set."""
    required_vars = {
        "JIRA_URL": "JIRA URL",
        "JIRA_USERNAME": "JIRA Username",
        "JIRA_PASSWORD": "JIRA Password",
        "ISSUE_ID": "Comma-separated list of Issue IDs",
        "BATCH_SIZE": "Number of issues to process"
    }
    for var, description in required_vars.items():
        if not os.getenv(var):
            print(f"Error: {description} ({var}) is not set in the environment variables.")
            sys.exit(1)


# Load environment variables from a .env file
load_dotenv()
# Jira credentials and URL from environment variables
JIRA_URL = os.getenv("JIRA_URL")
USERNAME = os.getenv("JIRA_USERNAME")
PASSWORD = os.getenv("JIRA_PASSWORD")
ISSUE_IDS = os.getenv("ISSUE_ID").split(',')  # Support multiple issue IDs
# Disable warnings for unverified HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# Define the IST timezone
ist = pytz.timezone('Asia/Kolkata')


def login_to_jira(username, password):
    """Log in to Jira and return the session cookies."""
    url = f"{JIRA_URL.rstrip('/')}/rest/auth/1/session"
    payload = {"username": username, "password": password}
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(url, json=payload, headers=headers, verify=False, timeout=10)
        response.raise_for_status()
    except RequestException as error:
        print(f"Login failed. Please check your Jira credentials and URL. Error: {error}")
        sys.exit(1)
    print(f"Successfully logged in as {username}")
    return response.cookies


def fetch_issue_with_changelog(issue_id, cookies):
    """Fetch issue details including changelog from Jira."""
    url = f"{JIRA_URL.rstrip('/')}/rest/api/2/issue/{issue_id}?expand=changelog"
    try:
        response = requests.get(url, cookies=cookies, verify=False, timeout=10)
        response.raise_for_status()
    except RequestException as error:
        print(f"Failed to fetch issue data for ID {issue_id}. Error: {error}")
        return None
    return response.json()


def fetch_time_tracking(issue_id):
    """Fetch estimated, remaining, and logged effort for the given Jira issue."""
    url = f"{JIRA_URL}/rest/api/2/issue/{issue_id}?fields=timetracking"
    response = requests.get(url, auth=(USERNAME, PASSWORD),
                            headers={"Accept": "application/json"}, verify=False,timeout=10)
    if response.status_code == 200:
        data = response.json()
        timetracking = data.get('fields', {}).get('timetracking', {})
        worklogs = data.get('fields', {}).get('worklog', {}).get('worklogs', [])
        return {
            "estimated": timetracking.get('originalEstimateSeconds', 0) / 3600,
            "remaining": timetracking.get('remainingEstimateSeconds', 0) / 3600,
            "logged": timetracking.get('timeSpentSeconds', 0) / 3600,
            "worklogs": worklogs
        }
    print(f"Failed to fetch time tracking for {issue_id}.Status code:{response.status_code}")
    return {"estimated": 0, "remaining": 0, "logged": 0, "worklogs": []}


def fetch_status_order(issue_id, cookies):
    """Fetch the status order for a Jira issue, including backward transitions."""
    url = f"{JIRA_URL.rstrip('/')}/rest/api/2/issue/{issue_id}?expand=changelog"
    try:
        response = requests.get(url, cookies=cookies, verify=False, timeout=10)
        response.raise_for_status()
    except RequestException as error:
        print(f"Failed to fetch status order for issue ID {issue_id}. Error: {error}")
        return []
    issue_data = response.json()
    if 'changelog' not in issue_data or 'histories' not in issue_data['changelog']:
        print(f"No changelog found for issue ID {issue_id}.")
        return []
    transitions = issue_data['changelog']['histories']
    status_order = []
    for history in transitions:
        for item in history['items']:
            if item['field'] == 'status':
                to_status = item.get('toString', None)
                if to_status and (not status_order or status_order[-1] != to_status):
                    status_order.append(to_status)
    return status_order


def fetch_issue_type_and_sprint(issue_id, cookies):
    """Fetch the issue type and sprint names from Jira for a given issue."""
    url = f"{JIRA_URL.rstrip('/')}/rest/api/2/issue/{issue_id}"
    try:
        response = requests.get(url, cookies=cookies, verify=False, timeout=10)
        response.raise_for_status()
    except RequestException as error:
        print(f"Failed to fetch issue data for ID {issue_id}. Error: {error}")
        return None, None  # Return None if error occurs
    # Log the entire JSON response for debugging
    try:
        issue_data = response.json()
    except ValueError as e:
        print(f"Failed to parse JSON response for issue {issue_id}: {e}")
        return None, None
    # Ensure 'fields' exists in the response
    if 'fields' not in issue_data:
        print(f"Error: 'fields' key is missing in the response for issue {issue_id}")
        return None, None

    # Extract the issue type (e.g., "Bug", "Story") from the issue's data
    issue_type = issue_data['fields'].get('issuetype', {}).get('name', 'N/A')
    if not issue_type:
        print(f"Issue type not found for issue {issue_id}")
    # Extract the sprint field (custom field 'customfield_10583')
    sprint_names = []  # List to store sprint names
    sprint_field = issue_data['fields'].get('customfield_10583', None)
    # Check if sprint_field exists and handle different types
    if sprint_field:
        if isinstance(sprint_field, list):  # If it's a list
            for sprint_data in sprint_field:
                # Extract sprint name from the string (format like 'name=SprintM 2024-05-22')
                if isinstance(sprint_data, str):
                    sprint_name = sprint_data.split('name=')[1].split(',')[0]
                    sprint_names.append(sprint_name)
        elif isinstance(sprint_field, dict):  # If it's a dictionary (edge case)
            sprint_name = sprint_field.get('name', 'N/A')  # Fetching full name from the dictionary
            sprint_names.append(sprint_name)
        elif isinstance(sprint_field, str):  # If it's a string
            sprint_names.append(sprint_field)  # Directly use it as the sprint name
    else:
        print(f"Sprint field not found for issue {issue_id}")
    # Join all sprint names into a single string (separated by commas) if multiple sprints are found
    sprint_names_str = ', '.join(sprint_names) if sprint_names else 'No Sprint'
    return issue_type, sprint_names_str


def calculate_durations(issue_data):
    """Calculate durations between status changes."""
    changelog = issue_data.get('changelog', {})
    histories = changelog.get('histories', [])
    if not histories:
        print("No changelog data available.")
        return []
    status_changes = [
        (item.get('fromString'), item.get('toString'), history.get('created'))
        for history in histories
        for item in history.get('items', [])
        if item.get('field') == 'status'
    ]
    if not status_changes:
        print("No status changes found in changelog.")
        return []
    durations = []
    last_timestamp = datetime.strptime(issue_data['fields']['created'], '%Y-%m-%dT%H:%M:%S.%f%z')
    for prev_status, curr_status, curr_time in status_changes:
        try:
            curr_time_dt = datetime.strptime(curr_time, '%Y-%m-%dT%H:%M:%S.%f%z')
            durations.append({
                "from": prev_status,
                "to": curr_status,
                "duration": curr_time_dt - last_timestamp
            })
            last_timestamp = curr_time_dt
        except ValueError as error:
            print(f"Error parsing date/time: {curr_time}. Error: {error}")
    return durations


def format_duration_days(duration):
    """Format a timedelta into years,months,weeks,days,hours,minutes,and seconds,excluding zeros."""
    total_seconds = int(duration.total_seconds())
    # Define constants for time conversions
    seconds_per_minute = 60
    seconds_per_hour = seconds_per_minute * 60
    seconds_per_day = seconds_per_hour * 24
    days = total_seconds // seconds_per_day
    total_seconds %= seconds_per_day
    hours = total_seconds // seconds_per_hour
    total_seconds %= seconds_per_hour
    minutes = total_seconds // seconds_per_minute
    seconds = total_seconds % seconds_per_minute
    # Collect non-zero parts
    parts = []
    if days > 0:
        parts.append(f"{days} day{'s' if days > 1 else ''}")
    if hours > 0:
        parts.append(f"{hours} hour{'s' if hours > 1 else ''}")
    if minutes > 0:
        parts.append(f"{minutes} minute{'s' if minutes > 1 else ''}")
    if seconds > 0:
        parts.append(f"{seconds} second{'s' if seconds > 1 else ''}")
    return ', '.join(parts) if parts else '0 seconds'


def format_duration(duration):
    """Format a timedelta into years, months, weeks, days, hours, minutes,
      and seconds, excluding zeros."""
    total_seconds = int(duration.total_seconds())
    # Define constants for time conversions
    time_units = [
        ("year", 365 * 24 * 60 * 60),  # seconds in a year
        ("month", 30 * 24 * 60 * 60),  # seconds in a month (approx)
        ("week", 7 * 24 * 60 * 60),    # seconds in a week
        ("day", 24 * 60 * 60),          # seconds in a day
        ("hour", 60 * 60),              # seconds in an hour
        ("minute", 60),                  # seconds in a minute
        ("second", 1)                   # seconds in a second
    ]
    parts = []
    for unit, seconds_in_unit in time_units:
        if total_seconds >= seconds_in_unit:
            count = total_seconds // seconds_in_unit
            total_seconds %= seconds_in_unit
            parts.append(f"{count} {unit}{'s' if count > 1 else ''}")
    return ', '.join(parts) if parts else '0 seconds'


def fetch_all_worklogs(issue_key):
    """Fetch all worklogs for a given issue with pagination."""
    all_worklogs = []
    start_at = 0
    max_results = 1000
    while True:
        response = requests.get(
            f'{JIRA_URL}/rest/api/2/issue/{issue_key}/worklog',
            auth=HTTPBasicAuth(USERNAME, PASSWORD),
            verify=False,timeout=10,  # Disable SSL verification
            params={'startAt': start_at, 'maxResults': max_results}
        )
        data = response.json()
        all_worklogs.extend(data.get('worklogs', []))
        if start_at + max_results >= data.get('total', 0):
            break
        start_at += max_results
    return all_worklogs

def accumulate_effort_per_status(issue_data):
    """Accumulate effort spent on each status from worklogs, ensuring all statuses are included."""
    time_spent_per_status = {}
    changelog = issue_data.get('changelog', {}).get('histories', [])
    # Extract changelog entries to track status changes
    status_change_dates = []
    initial_status = None
    for entry in changelog:
        for item in entry.get('items', []):
            if item['field'] == 'status':
                status_change_dates.append((item['toString'], entry['created']))
                if initial_status is None:
                    initial_status = item['fromString']
    # Insert initial status if it exists
    if initial_status and changelog:
        status_change_dates.insert(0, (initial_status, changelog[0]['created']))
    # Fetch all worklogs for the issue and initialize time spent per status
    worklogs = fetch_all_worklogs(issue_data['key'])
    for status, _ in status_change_dates:
        time_spent_per_status[status] = 0
    # Iterate through each worklog and determine its associated status
    for worklog in worklogs:
        time_spent_hours = worklog.get('timeSpentSeconds', 0) / 3600
        worklog_date = worklog['created']
        # Find the current status at the time the work was logged
        current_status = initial_status
        for status, change_date in reversed(status_change_dates):
            if worklog_date >= change_date:
                current_status = status
                break
        if current_status:
            time_spent_per_status[current_status] += time_spent_hours
    return time_spent_per_status


def get_issue_ids():
    """Retrieve issue IDs from environment variable or user input."""
    issue_ids_env = os.getenv("ISSUE_ID")
    if issue_ids_env:
        issue_ids = [issue_id.strip() for issue_id in issue_ids_env.split(',')]
        if not issue_ids:
            print("No issue IDs provided in the environment variable.")
            sys.exit(1)
    else:
        issue_ids = input("Enter the issue IDs separated by commas: ").strip().split(',')
        if not issue_ids:
            print("No issue IDs provided.")
            sys.exit(1)
    # Sort the issue IDs numerically based on the number part
    sorted_issue_ids = sorted(issue_ids, key=lambda x: int(re.search(r'(\d+)', x).group()))
    print(f"Using sorted issue IDs: {sorted_issue_ids}")
    return sorted_issue_ids


def fetch_creation_and_last_transition_dates(issue_data):
    """Fetch the creation date and last transition date for the issue."""
    # Fetch creation date
    creation_date = issue_data['fields']['created']
    last_transition_date = None
    # Check if changelog and histories exist
    if 'changelog' in issue_data and 'histories' in issue_data['changelog']:
        transitions = issue_data['changelog']['histories']
        # Iterate over transitions to find the last one
        for upd in transitions:
            for item in upd["items"]:
                if item["field"] == "status":
                    last_transition_date = upd["created"]
    return creation_date, last_transition_date


def format_detailed_duration(hours):
    """Format time in days, hours, minutes, seconds"""
    total_seconds = int(hours * 3600)  # Convert hours to seconds
    # Calculate the components
    days = total_seconds // (6 * 3600)  # 6 work hours in a day
    total_seconds %= (6 * 3600)
    hours = total_seconds // 3600
    total_seconds %= 3600
    minutes = total_seconds // 60
    seconds = total_seconds % 60
    parts = []
    if days > 0:
        parts.append(f"{days} days")
    if hours > 0:
        parts.append(f"{hours} hours")
    if minutes > 0:
        parts.append(f"{minutes} minutes")
    if seconds > 0:
        parts.append(f"{seconds} seconds")
    return ', '.join(parts) if parts else '0 seconds'


def initialize_workbook():
    """Initialize the workbook and create necessary sheets."""
    workbook = Workbook()
    # Create the main reporting sheet
    report_sheet = workbook.active
    report_sheet.title = "State Transition Summary"
    report_sheet.append([
        "Issue ID",
        "Issue Type",
        "Sprint Name",
        "Creation Date",
        "Last Transition Date",
        "From State",
        "To State",
        "Transition Flow",
        "Transition Duration",
        "Transition Duration(Days)",
        "Transition Duration(Hours)",
        "Transitions_Count"
        ])
    detailed_report_sheet = workbook.create_sheet(title="State Transition Details")
    detailed_report_sheet.append([
        "Issue ID",
        "Creation Date",
        "Last Transition Date",
        "From State",
        "To State",
        "Detailed Transition Flow",
        "Duration",
        "Duration(Days)"
    ])
    # Create a sheet for efforts
    efforts_sheet = workbook.create_sheet(title="Issues Summary")
    efforts_sheet.append([
        "Issue ID","Issue Type","Sprint Name", "Estimated Effort (h)", "Remaining Effort (h)",
        "Logged Effort (h)","Rounded Logged Effort(h)",
        "Logged Effort(Days)","Variance","Total Days Spent", "Start Date", "End Date"
    ])
    # Create a sheet for time spent per status
    time_spent_sheet = workbook.create_sheet(title="State Effort and Duration")
    time_spent_sheet.append(["Issue ID","Issue Type","Sprint Name", "Status",
                             "State Logged Effort(Hours)",
                             "State Rounded Effort(Hours)",
                             "Detailed Logged Effort","State Logged Effort(Days)"])
     # Return a dictionary with the workbook and its sheets
    return {
        "workbook": workbook,
        "report_sheet": report_sheet,
        "detailed_report_sheet": detailed_report_sheet,
        "efforts_sheet": efforts_sheet,
        "time_spent_sheet": time_spent_sheet
    }


def process_issue(issue_id, cookies, sheets):
    """Process a single Jira issue and append data to the relevant sheets."""
    issue_data = fetch_issue_with_changelog(issue_id, cookies)
    if not issue_data:
        print(f"No data found for issue ID: {issue_id}")
        return
    time_tracking = fetch_time_tracking(issue_id)
    creation_date, last_transition_date = fetch_creation_and_last_transition_dates(issue_data)
    # Convert creation_date and last_transition_date to IST
    creation_dt = parser.isoparse(creation_date).astimezone(ist)
    last_transition_dt = (
        parser.isoparse(last_transition_date).astimezone(ist)
        if last_transition_date
        else None
    )
    # Fetch issue type and sprint name
    issue_type, sprint_name = fetch_issue_type_and_sprint(issue_id, cookies)
    # Calculate total days spent
    total_days_spent = (last_transition_dt - creation_dt).days if last_transition_dt else 0
    # Prepare issue data dictionary
    issue_info = {
        "issue_id": issue_id,
        "time_tracking": time_tracking,
        "total_days_spent": total_days_spent,
        "creation_dt": creation_dt,
        "last_transition_dt": last_transition_dt,
        "durations": calculate_durations(issue_data)  # Add durations here
    }
    # Append data to the sheets using the issue_info dictionary
    append_report_data(sheets['report_sheet'], sheets['detailed_report_sheet'],
                       issue_info,issue_type, sprint_name)
    append_effort_data(sheets['efforts_sheet'], issue_info,issue_type, sprint_name)
    append_time_spent_data(sheets['time_spent_sheet'], issue_id,
                           accumulate_effort_per_status(issue_data),issue_type,
        sprint_name,)
    # Print summary to console
    print_summary(issue_id, time_tracking, total_days_spent)


def append_report_data(report_sheet, detailed_report_sheet, issue_data,issue_type, sprint_name):
    """Append issue data to the report sheet."""
    issue_id = issue_data['issue_id']
    creation_dt = issue_data['creation_dt']
    last_transition_dt = issue_data['last_transition_dt']
    durations = issue_data['durations']

    transition_data = calculate_transition_data(durations)

    append_summary_report(report_sheet, issue_id, creation_dt, last_transition_dt,
                          transition_data,issue_type, sprint_name)
    append_detailed_report(detailed_report_sheet,issue_id,creation_dt,last_transition_dt,durations)


def calculate_data(transition_durations, transition_execution_times):
    """Process the transition durations and execution times."""
    transition_info = []
    for (from_state, to_state), total_duration in transition_durations.items():
        days, total_hours = calculate_duration_components(total_duration)
        execution_times = transition_execution_times.get((from_state, to_state), 0)
        formatted_transition_str = f"{from_state} -> {to_state}"
        transition_info.append({
            'from_state': from_state,
            'to_state': to_state,
            'formatted_transition_str': formatted_transition_str,
            'total_duration': total_duration,
            'days': days,
            'total_hours': total_hours,
            'execution_times': execution_times
        })
    return transition_info


#pylint: disable=too-many-arguments
def append_summary_report(report_sheet, issue_id, creation_dt, last_transition_dt,
                          transition_data, issue_type, sprint_name):
    """Append summary data for transitions to the report sheet."""
    transition_durations, transition_execution_times = transition_data
    # Pass transition_durations and transition_execution_times as separate arguments
    transition_info = calculate_data(transition_durations, transition_execution_times)

    # Append each transition's information to the report sheet
    for info in transition_info:
        report_sheet.append([
            issue_id,
            issue_type,  # Added issue_type here
            sprint_name,
            creation_dt.strftime('%Y-%m-%d %H:%M:%S'),
            last_transition_dt.strftime('%Y-%m-%d %H:%M:%S') if last_transition_dt else None,
            info['from_state'],
            info['to_state'],
            info['formatted_transition_str'],
            format_duration_days(info['total_duration']),
            info['days'],
            info['total_hours'],
            info['execution_times'],
        ])


def append_detailed_report(detailed_report_sheet,issue_id,creation_dt,last_transition_dt,durations):
    """Append detailed transition data to the report sheet."""
    for transition in durations:
        days, _ = calculate_duration_components(transition["duration"])
        detailed_report_sheet.append([
            issue_id,
            creation_dt.strftime('%Y-%m-%d %H:%M:%S'),
            last_transition_dt.strftime('%Y-%m-%d %H:%M:%S') if last_transition_dt else None,
            transition["from"],
            transition["to"],
            f"{transition['from']} -> {transition['to']}",
            format_duration(transition["duration"]),
            days,
        ])


def calculate_transition_data(durations):
    """Calculate total durations and execution times for transitions."""
    transition_durations = {}
    transition_execution_times = {}

    for transition in durations:
        transition_key = (transition["from"], transition["to"])
        duration = transition["duration"]
        if transition_key not in transition_durations:
            transition_durations[transition_key] = duration
            transition_execution_times[transition_key] = 1
        else:
            transition_durations[transition_key] += duration
            transition_execution_times[transition_key] += 1

    return transition_durations, transition_execution_times


def calculate_duration_components(total_duration):
    """Calculate rounded days and total hours from a duration."""
    total_seconds = total_duration.total_seconds()
    days = total_seconds // (24 * 3600)
    remaining_seconds = total_seconds % (24 * 3600)
    remaining_hours = remaining_seconds // 3600

    # Calculate rounded days
    if total_seconds > 0:
        if days == 0 and remaining_seconds > 0:
            # Case 1: If no full days but there are hours or minutes
            rounded_days = 1
        elif days > 0 and remaining_hours >= 12:
            # Case 2: If there are full days and more than half a day in hours
            rounded_days = days + 1
        else:
            # Case 3: Otherwise, keep the calculated days
            rounded_days = days
    else:
        # If total_duration is zero or negative
        rounded_days = 0

    total_hours = max(1, int(total_seconds // 3600)) if total_seconds > 0 else 0
    return rounded_days, total_hours


def format_hours_minutes(total_hours):
    """Format total hours into hours and minutes, showing only non-zero values."""
    hours = int(total_hours)
    minutes = int((total_hours - hours) * 60)
    parts = []
    if hours > 0:
        parts.append(f"{hours}h")
    if minutes > 0:
        parts.append(f"{minutes}m")
    return ' '.join(parts) if parts else '0m'


def append_effort_data(efforts_sheet, issue_info,issue_type, sprint_name):
    """Append effort data to the efforts sheet, reflecting accurate time tracking values."""
    # Extract time tracking values
    estimated_effort = issue_info["time_tracking"]["estimated"]
    remaining_effort = issue_info["time_tracking"]["remaining"]
    logged_effort = issue_info["time_tracking"]["logged"]
    # Format logged effort for display
    logged_effort_formatted = format_hours_minutes(logged_effort)
    # Rounding logic
    if logged_effort - math.floor(logged_effort) >= 0.5:
        logged_effort_rounded = math.ceil(logged_effort)
    else:
        logged_effort_rounded = math.floor(logged_effort)
    # Calculate logged effort in days (1 day = 6 hours)
    logged_effort_days = logged_effort / 6
    logged_effort_days_rounded = round(logged_effort_days)
    # Calculate variance
    if estimated_effort > 0:
        variance = ((logged_effort_rounded - estimated_effort) /(estimated_effort)) * 100
        variance = round(variance)
    else:
        variance = 0  # Handle division by zero if needed
    # Append the data to the efforts sheet
    efforts_sheet.append([
        issue_info["issue_id"],
        issue_type,
        sprint_name,
        estimated_effort,
        remaining_effort,
        logged_effort_formatted,
        logged_effort_rounded,
        logged_effort_days_rounded,
        variance,
        issue_info["total_days_spent"],
        issue_info["creation_dt"].strftime('%Y-%m-%d %H:%M:%S'),
        issue_info["last_transition_dt"].strftime('%Y-%m-%d %H:%M:%S')
        if issue_info["last_transition_dt"] else None
    ])


def calculate_days_from_hours(hours):
    """Calculate the total days from hours and return as an integer."""
    return round(hours / 6)


def append_time_spent_data(time_spent_sheet, issue_id, time_spent_per_status,issue_type,
        sprint_name):
    """Append time spent data per status to the time spent sheet."""
    for status, hours in time_spent_per_status.items():
        detailed_time_spent = format_detailed_duration(hours)
        rounded_hours = math.ceil(hours) if hours * 60 >= 30 else math.floor(hours)
        detailed_time_spent_days = calculate_days_from_hours(hours)
        time_spent_sheet.append([
            issue_id,
            issue_type,
            sprint_name,
            status,
            format_hours_minutes(hours),
            rounded_hours,
            detailed_time_spent,
            detailed_time_spent_days
        ])


def print_summary(issue_id, time_tracking, total_days_spent):
    """Print a summary of the issue data to the console."""
    print(f"Issue ID: {issue_id}")
    print(f"Estimated Effort: {time_tracking['estimated']} hours")
    print(f"Remaining Effort: {time_tracking['remaining']} hours")
    print(f"Logged Effort: {time_tracking['logged']} hours")
    print(f"Total Days Spent: {total_days_spent} days")
    print("-" * 40)  # Separator for clarity


def create_output_folder(script_directory):
    """
    Create an output folder for storing files.
    Args:
        script_directory (str): The directory where the script is located.
    Returns:
        str: The path to the created output folder.
    """
    output_folder = os.path.join(script_directory, 'Reports')
    os.makedirs(output_folder, exist_ok=True)
    return output_folder


def process_issues(issue_ids_to_process, cookies, sheets):
    """
    Process a list of issue IDs and populate the corresponding sheets in the workbook.
    Args:
        issue_ids_to_process (list): A list of issue IDs to process.
        cookies (dict): Authentication cookies for accessing the Jira API.
        report_sheet (Worksheet): The sheet in the workbook for the report.
        efforts_sheet (Worksheet): The sheet in the workbook for effort tracking.
        time_spent_sheet (Worksheet): The sheet in the workbook for time spent tracking.
    Returns:
        None
    """
    for issue_id in issue_ids_to_process:
        process_issue(issue_id, cookies, sheets)


def save_report(workbook, output_folder):
    """
    Save the workbook to the specified output folder.
    Args:
        workbook (Workbook): The workbook object to be saved.
        output_folder (str): The directory where the workbook should be saved.
    Returns:
        str: The path to the saved workbook file.
    """
    output_file = os.path.join(output_folder, "Jira_report.xlsx")
    workbook.save(output_file)
    print(f"Report saved as {output_file}")
    return output_file


def convert_excel_to_json(output_file, output_folder):
    """
    Convert an Excel file to JSON format using an external script.
    This function calls the 'excel_to_json.py' script to convert the
    Excel file into JSON files, which are saved in the specified output folder.
    Args:
        output_file (str): The path to the Excel file that needs to be converted.
        output_folder (str): The directory where the JSON files should be saved.
    Returns:
        None: This function does not return a value.
    """
    command = ["python", "ej10aa.py", output_file, output_folder]
    try:
        subprocess.run(command, check=True)
        print(f"Jira data successfully processed and saved to JSON in {output_folder}.")
    except subprocess.CalledProcessError as error:
        print(f"Error while running the script: {error}")


def parse_arguments():
    """Parse command-line arguments and return them."""
    arg_parser = argparse.ArgumentParser(description="Environment Variables to be set:")
    arg_parser.add_argument('--JIRA_URL', type=str, help="JIRA URL.")
    arg_parser.add_argument('--JIRA_USERNAME', type=str, help="Your JIRA User ID.")
    arg_parser.add_argument('--JIRA_PASSWORD', type=str, help="Your JIRA Password.")
    arg_parser.add_argument('--ISSUE_ID', type=str, help="Comma-separated list of issue IDs.")
    arg_parser.add_argument('--BATCH_SIZE',type=int,help="Number of issues to process.")
    return arg_parser.parse_args()


def set_environment_variables(args):
    """Set environment variables from command line arguments if provided."""
    if args.JIRA_URL:
        os.environ["JIRA_URL"] = args.JIRA_URL
    if args.JIRA_USERNAME:
        os.environ["JIRA_USERNAME"] = args.JIRA_USERNAME
    if args.JIRA_PASSWORD:
        os.environ["JIRA_PASSWORD"] = args.JIRA_PASSWORD
    if args.ISSUE_ID:
        os.environ["ISSUE_ID"] = args.ISSUE_ID
    if args.BATCH_SIZE:
        os.environ["BATCH_SIZE"] = str(args.BATCH_SIZE)


def main():
    """Main function to execute the reporting script."""
    # Validate environment variables before proceeding
    validate_environment_variables()
    script_directory = os.path.dirname(os.path.abspath(__file__))
    output_folder = create_output_folder(script_directory)
    args = parse_arguments()
    set_environment_variables(args)
    issue_ids = get_issue_ids()
    batch_size = int(os.getenv("BATCH_SIZE", len(issue_ids)))# Default to length if not set
    batch_size = min(batch_size, len(issue_ids))
    issue_ids_to_process = issue_ids[:batch_size]
    if not issue_ids_to_process:
        print("No issue IDs available to process.")
        return
    cookies = login_to_jira(USERNAME, PASSWORD)
    sheets = initialize_workbook()
    # Access the workbook and sheets using the dictionary
    workbook = sheets["workbook"]
    print(f"Processing the following issue IDs: {issue_ids_to_process}")
    process_issues(issue_ids_to_process,cookies,sheets)
    output_file = save_report(workbook, output_folder)
    convert_excel_to_json(output_file, output_folder)


if __name__ == "__main__":
    main()
